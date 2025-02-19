"""
import csv
import openpyxl

# Iterator 패턴 적용: Excel 데이터를 순회하는 반복자 클래스
class ExcelIterator:
    def __init__(self, source_file):
        self.wb = openpyxl.load_workbook(source_file, read_only=True)
        self.sheet = self.wb.active
        self.rows = iter(self.sheet.iter_rows(values_only=True))  # 값만 가져오기

    def __iter__(self):
        return self

    def __next__(self):
        try:
            return next(self.rows)  # 다음 행 반환
        except StopIteration:
            self.wb.close()
            raise

# 클라이언트 코드: 반복자를 사용하여 데이터를 CSV에 추가하는 클래스
class DataExporter:
    def __init__(self, source, target):
        self.source = source
        self.target = target

    def append_to_csv(self):
        iterator = ExcelIterator(self.source)  # 반복자 생성

        with open(self.target, mode='a', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)
            
            for row in iterator:  # 반복자를 통해 데이터 순회
                writer.writerow(row)

# 사용 예시
source_file = ""
target_csv = ""

exporter = DataExporter(source_file, target_csv)
exporter.append_to_csv()
"""



"""
import csv
import openpyxl

# 제너레이터 패턴 적용: Excel 데이터를 순회하는 제너레이터 함수
def excel_iterator(source_file):
    wb = openpyxl.load_workbook(source_file, read_only=True)
    sheet = wb.active
    for row in sheet.iter_rows(values_only=True):  # 값만 가져오기
        yield row  # 행을 하나씩 반환
    wb.close()

# 클라이언트 코드: 제너레이터를 사용하여 데이터를 CSV에 추가하는 클래스
class DataExporter:
    def __init__(self, source, target):
        self.source = source
        self.target = target

    def append_to_csv(self):
        # 제너레이터 함수 호출
        iterator = excel_iterator(self.source)

        with open(self.target, mode='a', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)
            
            for row in iterator:  # 제너레이터를 통해 데이터 순회
                writer.writerow(row)

# 사용 예시
source_file = ""
target_csv = ""

exporter = DataExporter(source_file, target_csv)
exporter.append_to_csv()
"""